import os
import argparse
from openpyxl import Workbook, utils
from PIL import Image as PILImage
from openpyxl.drawing.image import Image

def resize_images(directory, target_height=150):
    """
    Resize images in the specified directory to the target height.

    input: path to directory containing images
    target_height: resize images to this value to easily fit in the sheet
    """
    for filename in os.listdir(directory):
        if filename.lower().endswith(('.png', '.jpg', '.jpeg')):
            file_path = os.path.join(directory, filename)
            img = PILImage.open(file_path)

            hpercent = (target_height / float(img.size[1]))
            wsize = int((float(img.size[0]) * float(hpercent)))

            img_resized = img.resize((wsize, target_height), PILImage.LANCZOS)
            img_resized.save(file_path)

def create_workbook(directory, workbook_name):
    """Create an Excel workbook with image names and resized images."""
    wb = Workbook()
    ws = wb.active

    for idx, filename in enumerate(os.listdir(directory)):
        if filename.lower().endswith(('.jpg', '.png')):
            ws.cell(row=idx+1, column=1, value=os.path.splitext(filename)[0])

            img_path = os.path.join(directory, filename)
            img = Image(img_path)
            ws.add_image(img, f'B{idx+1}')
            ws.row_dimensions[idx+1].height = 130 # NOTE: CHANGE THIS FOR CHANGING ROW HEIGHT
    
    # changing column A width to fit content
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        adjusted_width = (length + 2)  # Adding a little extra width for padding
        ws.column_dimensions[utils.get_column_letter(column_cells[0].column)].width = adjusted_width

    ws.column_dimensions['B'].width = 25 # NOTE: CHANGE THIS FOR CHANGING ROW WIDTH OF IMAGE COLUMN
    wb.save(workbook_name)

def main(directory_path, workbook_name, target_height):
    resize_images(directory_path, target_height)
    # print("All images have been resized.")
    create_workbook(directory_path, workbook_name)
    print(f"Workbook '{workbook_name}' has been created with the images.")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Resize images and create an Excel workbook.")
    parser.add_argument("directory_path", type=str, help="Directory path containing the images")
    parser.add_argument("workbook_name", type=str, help="Name of the workbook to save")
    parser.add_argument("--target_height", type=int, default=150, help="Target height for the resized images in pixels")

    args = parser.parse_args()

    main(args.directory_path, args.workbook_name, args.target_height)